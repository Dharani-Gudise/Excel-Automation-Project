import openpyxl
from openpyxl.chart.reference import Reference
from openpyxl.chart.bar_chart import BarChart
from openpyxl.drawing.colors import ColorChoice
from openpyxl.drawing.fill import PatternFillProperties


def process_workbook(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb["Sheet1"]
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        cell1 = sheet.cell(row, 4)
        cell1.value = corrected_price
    values = Reference(sheet, min_row=2, max_row=4, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "e2")
    chart.title = "Tranactions"
    chart.x_axis.title = "corrected prices"
    sequence = chart.series[0]
    my_excel = PatternFillProperties(prst="cross")
    my_excel.foreground = ColorChoice(prstClr="blue")
    my_excel.background = ColorChoice(prstClr="red")
    sequence.graphicalProperties.pattFill = my_excel
    wb.save(filename)


process_workbook("C:\\Users\\Gudise Dharani\\Downloads\\Excel Files\\transactions.xlsx")
